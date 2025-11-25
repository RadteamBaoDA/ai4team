"""
Docling-enhanced converter for advanced layout recognition
Optimized for Excel files with complex table structures
"""

from pathlib import Path
from datetime import datetime
from typing import Optional

from .base_converter import BaseConverter
from ..utils import setup_logger
from config.settings import (
    CONVERSION_TIMEOUT,
    RAG_OPTIMIZATION_ENABLED,
    EXCEL_TABLE_MAX_ROWS_PER_PAGE,
    EXCEL_TABLE_OPTIMIZATION,
    EXCEL_PRINT_ROW_COL_HEADERS,
    EXCEL_PRINT_GRIDLINES,
    EXCEL_ADD_CITATION_HEADERS,
    CITATION_INCLUDE_FILENAME,
    CITATION_INCLUDE_PAGE,
    CITATION_INCLUDE_DATE,
    CITATION_DATE_FORMAT,
    DOCLING_API_ENABLED,
    DOCLING_API_URL,
    DOCLING_API_TIMEOUT,
)


class DoclingConverter(BaseConverter):
    """
    Converts documents using Docling for enhanced layout recognition.
    
    Best for:
    - Excel files with complex tables (merged cells, borderless, rotated)
    - Multi-sheet workbooks requiring semantic understanding
    - Documents needing advanced layout analysis
    
    Requires: pip install docling reportlab pypdf
    """
    
    def __init__(self):
        self.logger = setup_logger(__name__)
        self._docling_available = False
        self._unicode_font = None
        self._unicode_font_bold = None
        self._font_map = {
            'japanese': None,
            'korean': None,
            'chinese': None,
            'vietnamese': None,
            'generic': None
        }
        self._check_availability()
        if self._docling_available:
            self._register_unicode_fonts()
    
    def _check_availability(self):
        """Check if Docling and dependencies are available"""
        try:
            from reportlab.lib.pagesizes import letter  # noqa: F401
            
            # If API is enabled, we don't strictly need the local docling package
            if DOCLING_API_ENABLED:
                self._docling_available = True
                self.logger.debug(f"Docling API enabled (URL: {DOCLING_API_URL})")
                return

            import docling  # noqa: F401
            self._docling_available = True
            self.logger.debug("Docling converter available")
        except ImportError as e:
            self.logger.debug(f"Docling/ReportLab not available: {e}")
            self._docling_available = False
    
    def _register_unicode_fonts(self):
        """Register Unicode-capable fonts for CJK/Japanese support"""
        try:
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            from pathlib import Path
            
            # Try to find and register Unicode fonts
            # Note: Single font cannot support all CJK languages perfectly
            # MS-Gothic: Japanese + some Chinese ✓, Korean ✗
            # Malgun: Korean ✓, Japanese/Chinese limited
            # SimSun: Chinese ✓, Japanese/Korean limited
            # Best practice: Use language-specific files or install NotoSansCJK
            font_candidates = [
                # Universal / Wide coverage fonts
                ('c:/windows/fonts/arialuni.ttf', 'Arial-Unicode'), # Best if available
                ('c:/windows/fonts/micross.ttf', 'MicrosoftSansSerif'), # Good fallback
                ('c:/windows/fonts/tahoma.ttf', 'Tahoma'),          # Good for Vietnamese + Latin
                ('c:/windows/fonts/segoeui.ttf', 'SegoeUI'),        # Windows System Font
                
                # Japanese fonts (best for Japanese + partial Chinese)
                ('c:/windows/fonts/msgothic.ttc', 'MS-Gothic'),
                ('c:/windows/fonts/msmincho.ttc', 'MS-Mincho'),
                ('c:/windows/fonts/meiryo.ttc', 'Meiryo'),
                ('c:/windows/fonts/yumin.ttf', 'YuMincho'),
                ('c:/windows/fonts/yumindb.ttf', 'YuMincho-Bold'),
                # Korean fonts
                ('c:/windows/fonts/malgun.ttf', 'Malgun'),
                ('c:/windows/fonts/malgunbd.ttf', 'Malgun-Bold'),
                # Chinese fonts
                ('c:/windows/fonts/simsun.ttc', 'SimSun'),
                ('c:/windows/fonts/simhei.ttf', 'SimHei'),
                # Generic Unicode fonts
                ('c:/windows/fonts/arial.ttf', 'Arial-Unicode'),
                ('c:/windows/fonts/arialbd.ttf', 'Arial-Unicode-Bold'),
            ]
            
            registered_fonts = []
            registered_regular = False
            registered_bold = False
            
            for font_path, font_name in font_candidates:
                if not Path(font_path).exists():
                    continue
                    
                try:
                    # For .ttc files, use subfontIndex
                    if font_path.endswith('.ttc'):
                        font = TTFont(font_name, font_path, subfontIndex=0)
                    else:
                        font = TTFont(font_name, font_path)
                    
                    pdfmetrics.registerFont(font)
                    registered_fonts.append(font_name)
                    
                    # Map fonts to categories
                    lower_name = font_name.lower()
                    if 'gothic' in lower_name or 'mincho' in lower_name or 'meiryo' in lower_name:
                        if not self._font_map['japanese']: self._font_map['japanese'] = font_name
                    elif 'malgun' in lower_name:
                        if not self._font_map['korean']: self._font_map['korean'] = font_name
                    elif 'simsun' in lower_name or 'simhei' in lower_name:
                        if not self._font_map['chinese']: self._font_map['chinese'] = font_name
                    elif 'tahoma' in lower_name or 'segoe' in lower_name or 'sansserif' in lower_name:
                        if not self._font_map['vietnamese']: self._font_map['vietnamese'] = font_name
                        if not self._font_map['generic']: self._font_map['generic'] = font_name
                    
                    if not registered_regular:
                        self._unicode_font = font_name
                        registered_regular = True
                        self.logger.info(f"Registered primary Unicode font: {font_name}")
                    
                    if 'bold' in font_name.lower() and not registered_bold:
                        self._unicode_font_bold = font_name
                        registered_bold = True
                        self.logger.info(f"Registered Unicode bold font: {font_name}")
                    
                    # Don't break - register all available fonts for maximum coverage
                        
                except Exception as font_error:
                    self.logger.debug(f"Could not register font {font_name}: {font_error}")
                    continue
            
            if registered_fonts:
                self.logger.info(f"Total registered fonts: {len(registered_fonts)} ({', '.join(registered_fonts[:3])}...)")
            
            if not registered_regular:
                self.logger.warning("No Unicode font registered - CJK characters may not display correctly")
            else:
                # Set default bold font if not found
                if not registered_bold:
                    self._unicode_font_bold = self._unicode_font
                    
        except Exception as e:
            self.logger.warning(f"Unicode font registration failed: {e}")
            self._unicode_font = None
            self._unicode_font_bold = None
    
    def is_available(self) -> bool:
        """Check if Docling converter is available"""
        return self._docling_available
    
    def convert(self, input_file: Path, output_file: Path) -> bool:
        """
        Convert document using Docling for layout analysis + ReportLab for PDF generation
        
        Args:
            input_file: Source file (Excel, Word, PowerPoint, PDF, etc.)
            output_file: Destination PDF file
            
        Returns:
            True if successful, False otherwise
        """
        if not self.is_available():
            return False
        
        ext = input_file.suffix.lower()
        
        # Docling supports: PDF, DOCX, PPTX, XLSX, HTML, images, audio, video
        supported_formats = {
            '.xlsx', '.xls',  # Excel
            '.docx', '.doc',  # Word
            '.pptx', '.ppt',  # PowerPoint
            '.pdf',           # PDF (re-parsing)
            '.html', '.htm',  # HTML
            '.png', '.jpg', '.jpeg', '.tiff',  # Images
        }
        
        if ext not in supported_formats:
            return False
        
        try:
            if DOCLING_API_ENABLED:
                return self._convert_with_docling_api(input_file, output_file)
            return self._convert_with_docling(input_file, output_file)
        except Exception as e:
            self.logger.error(f"Docling conversion failed for {input_file.name}: {e}")
            return False
    
    def _convert_with_docling(self, input_file: Path, output_file: Path) -> bool:
        """Convert using Docling for layout analysis"""
        try:
            from docling.document_converter import DocumentConverter
            from reportlab.lib.pagesizes import letter, landscape
            from reportlab.lib import colors
            from reportlab.platypus import (
                SimpleDocTemplate,
                Table,
                TableStyle,
                Paragraph,
                PageBreak,
                Spacer,
            )
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.lib.units import inch
            
            self.logger.info(f"Converting {input_file.name} with Docling layout analysis")
            
            # Step 1: Analyze Excel file for frozen panes and table structure
            excel_metadata = {}
            if input_file.suffix.lower() in {'.xlsx', '.xls'}:
                excel_metadata = self._analyze_excel_file(input_file)
                if excel_metadata.get('sheet_count'):
                    self.logger.info(f"Excel analysis: {excel_metadata['sheet_count']} sheets, frozen panes: {len(excel_metadata.get('frozen_panes', {}))}")
            
            # Step 2: Parse document with Docling
            self.logger.info("Starting Docling document conversion...")
            converter = DocumentConverter()
            result = converter.convert(str(input_file))
            doc_obj = result.document
            
            self.logger.info(f"Docling conversion complete. Document type: {type(doc_obj).__name__}")
            self.logger.debug(f"Document attributes: {[attr for attr in dir(doc_obj) if not attr.startswith('_')][:20]}")
            
            # Step 3: Create PDF with ReportLab
            # Ensure output path is absolute and parent exists
            output_file = Path(output_file).resolve()
            output_file.parent.mkdir(parents=True, exist_ok=True)
            
            pdf_doc = SimpleDocTemplate(
                str(output_file),
                pagesize=landscape(letter) if self._is_wide_content(doc_obj) else letter,
                title=input_file.stem,
            )
            elements = []
            styles = getSampleStyleSheet()
            
            # Customize styles with Unicode fonts for CJK support
            if self._unicode_font:
                for style_name in styles.byName:
                    style = styles[style_name]
                    if hasattr(style, 'fontName'):
                        if 'Bold' in style.fontName or 'bold' in style_name.lower():
                            style.fontName = self._unicode_font_bold if self._unicode_font_bold else self._unicode_font
                        else:
                            style.fontName = self._unicode_font
                self.logger.debug(f"Applied Unicode font '{self._unicode_font}' to all paragraph styles")
            
            # Step 4: Add citation header if RAG enabled
            if RAG_OPTIMIZATION_ENABLED and EXCEL_ADD_CITATION_HEADERS:
                citation_parts = []
                if CITATION_INCLUDE_FILENAME:
                    citation_parts.append(f"Source: {input_file.name}")
                if excel_metadata.get('sheet_count'):
                    citation_parts.append(f"Sheets: {excel_metadata['sheet_count']}")
                if CITATION_INCLUDE_DATE:
                    citation_parts.append(f"Converted: {datetime.now().strftime(CITATION_DATE_FORMAT)}")
                
                if citation_parts:
                    header_text = " | ".join(citation_parts)
                    elements.append(Paragraph(f"<b>{header_text}</b>", styles['Normal']))
                    elements.append(Spacer(1, 0.2 * inch))
            
            # Step 5: Process Docling document structure with Excel metadata
            self._process_docling_document(doc_obj, elements, styles, input_file, excel_metadata)
            
            # Fallback: If no elements were added and this is an Excel file, use openpyxl directly
            if len(elements) <= 2 and input_file.suffix.lower() in {'.xlsx', '.xls'}:
                self.logger.warning("Docling extraction produced no content, using openpyxl fallback")
                self._add_excel_tables_with_openpyxl(input_file, elements, styles, excel_metadata)
            
            # Step 6: Build PDF
            if not elements:
                self.logger.error("No content to add to PDF!")
                elements.append(Paragraph(f"Error: No content could be extracted from {input_file.name}", styles['Normal']))
            
            pdf_doc.build(elements)
            
            # Step 7: Add metadata with Excel info
            if output_file.exists():
                self._apply_pdf_metadata(output_file, input_file, doc_obj, excel_metadata)
            
            self.logger.info(f"Successfully converted {input_file.name} with Docling (enhanced Excel mode)")
            return True
            
        except Exception as e:
            self.logger.error(f"Docling conversion error: {e}")
            return False
    
    def _convert_with_docling_api(self, input_file: Path, output_file: Path) -> bool:
        """Convert using Docling API"""
        try:
            import requests
            from reportlab.lib.pagesizes import letter, landscape
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.lib.units import inch
            
            self.logger.info(f"Converting {input_file.name} via Docling API at {DOCLING_API_URL}")
            
            # Step 1: Analyze Excel file for frozen panes and table structure
            excel_metadata = {}
            if input_file.suffix.lower() in {'.xlsx', '.xls'}:
                excel_metadata = self._analyze_excel_file(input_file)
                if excel_metadata.get('sheet_count'):
                    self.logger.info(f"Excel analysis: {excel_metadata['sheet_count']} sheets")
            
            # Step 2: Call API
            try:
                with open(input_file, 'rb') as f:
                    files = {'file': f}
                    # Try /convert endpoint first (standard for many wrappers)
                    api_url = DOCLING_API_URL.rstrip('/') + '/convert'
                    self.logger.debug(f"Posting to {api_url}")
                    
                    response = requests.post(
                        api_url, 
                        files=files, 
                        timeout=DOCLING_API_TIMEOUT
                    )
                    response.raise_for_status()
                    
                    # Parse JSON response
                    api_data = response.json()
                    self.logger.info("Received response from Docling API")
                    
            except Exception as api_error:
                self.logger.error(f"API request failed: {api_error}")
                return False
            
            # Step 3: Wrap API response to mimic Docling document
            class ApiItem:
                def __init__(self, data):
                    self._data = data
                    self.label = data.get('label', 'text')
                    self.type = data.get('type', self.label)
                    self.text = data.get('text', '')
                    self.content = data.get('content', self.text)
                    
                    # Handle table data
                    self.data = data.get('data', None)
                    
                    # Handle provenance/metadata
                    prov = data.get('prov', {})
                    self.prov = type('Prov', (), {'sheet_name': prov.get('sheet_name'), 'page_no': prov.get('page_no')})()
                    self.metadata = data.get('metadata', {})
                
                def export_to_dataframe(self):
                    # Helper for table extraction if data is in grid format
                    if self.data and 'grid' in self.data:
                        import pandas as pd
                        return pd.DataFrame(self.data['grid'])
                    return None

            class ApiDoc:
                def __init__(self, data):
                    self.data = data
                
                def iterate_items(self):
                    # Try to find items list in common locations
                    items = []
                    if isinstance(self.data, list):
                        items = self.data
                    elif isinstance(self.data, dict):
                        items = self.data.get('items', self.data.get('main-text', self.data.get('body', [])))
                    
                    for item in items:
                        yield ApiItem(item)
                
                def export_to_markdown(self):
                    return self.data.get('markdown', '') if isinstance(self.data, dict) else ''

            doc_obj = ApiDoc(api_data)
            
            # Step 4: Create PDF with ReportLab (Shared logic)
            output_file = Path(output_file).resolve()
            output_file.parent.mkdir(parents=True, exist_ok=True)
            
            pdf_doc = SimpleDocTemplate(
                str(output_file),
                pagesize=landscape(letter) if self._is_wide_content(doc_obj) else letter,
                title=input_file.stem,
            )
            elements = []
            styles = getSampleStyleSheet()
            
            # Customize styles with Unicode fonts
            if self._unicode_font:
                for style_name in styles.byName:
                    style = styles[style_name]
                    if hasattr(style, 'fontName'):
                        if 'Bold' in style.fontName or 'bold' in style_name.lower():
                            style.fontName = self._unicode_font_bold if self._unicode_font_bold else self._unicode_font
                        else:
                            style.fontName = self._unicode_font
            
            # Add citation header
            if RAG_OPTIMIZATION_ENABLED and EXCEL_ADD_CITATION_HEADERS:
                citation_parts = []
                if CITATION_INCLUDE_FILENAME:
                    citation_parts.append(f"Source: {input_file.name}")
                if excel_metadata.get('sheet_count'):
                    citation_parts.append(f"Sheets: {excel_metadata['sheet_count']}")
                if CITATION_INCLUDE_DATE:
                    citation_parts.append(f"Converted: {datetime.now().strftime(CITATION_DATE_FORMAT)}")
                
                if citation_parts:
                    header_text = " | ".join(citation_parts)
                    elements.append(Paragraph(f"<b>{header_text}</b>", styles['Normal']))
                    elements.append(Spacer(1, 0.2 * inch))
            
            # Process document structure
            self._process_docling_document(doc_obj, elements, styles, input_file, excel_metadata)
            
            # Fallback for Excel
            if len(elements) <= 2 and input_file.suffix.lower() in {'.xlsx', '.xls'}:
                self.logger.warning("API extraction produced no content, using openpyxl fallback")
                self._add_excel_tables_with_openpyxl(input_file, elements, styles, excel_metadata)
            
            # Build PDF
            if not elements:
                elements.append(Paragraph(f"Error: No content extracted from {input_file.name}", styles['Normal']))
            
            pdf_doc.build(elements)
            
            # Add metadata
            if output_file.exists():
                self._apply_pdf_metadata(output_file, input_file, doc_obj, excel_metadata)
            
            self.logger.info(f"Successfully converted {input_file.name} via Docling API")
            return True
            
        except Exception as e:
            self.logger.error(f"Docling API conversion error: {e}")
            return False
    
    def _detect_file_encoding(self, file_path: Path) -> str:
        """Detect file encoding for proper text reading"""
        try:
            import chardet
            with open(file_path, 'rb') as f:
                raw_data = f.read(10000)  # Read first 10KB
                result = chardet.detect(raw_data)
                encoding = result.get('encoding', 'utf-8')
                confidence = result.get('confidence', 0)
                self.logger.debug(f"Detected encoding: {encoding} (confidence: {confidence:.2%})")
                return encoding if confidence > 0.7 else 'utf-8'
        except ImportError:
            self.logger.debug("chardet not available, using utf-8")
            return 'utf-8'
        except Exception as e:
            self.logger.debug(f"Encoding detection failed: {e}, using utf-8")
            return 'utf-8'
    
    def _analyze_excel_file(self, input_file: Path) -> dict:
        """Analyze Excel file for frozen panes and table structure (like MS Office converter)"""
        metadata = {
            'frozen_panes': {},
            'sheet_count': 0,
            'sheet_names': [],
            'has_tables': False
        }
        
        try:
            import openpyxl
            workbook = openpyxl.load_workbook(str(input_file), read_only=True, data_only=True)
            metadata['sheet_count'] = len(workbook.worksheets)
            
            self.logger.debug(f"Analyzing {metadata['sheet_count']} worksheets")
            
            for sheet in workbook.worksheets:
                sheet_name = sheet.title
                metadata['sheet_names'].append(sheet_name)
                self.logger.debug(f"Found sheet: {sheet_name}")
                
                # Detect frozen panes (for header row repetition)
                try:
                    if sheet.freeze_panes:
                        cell = sheet.freeze_panes
                        if isinstance(cell, str):
                            # Parse cell reference like "A2" -> row 2
                            from openpyxl.utils import coordinate_from_string
                            col, row = coordinate_from_string(cell)
                            frozen_rows = row - 1  # Row 2 means 1 header row
                        else:
                            frozen_rows = cell.row - 1 if hasattr(cell, 'row') else 0
                        
                        if frozen_rows > 0:
                            metadata['frozen_panes'][sheet_name] = frozen_rows
                            self.logger.info(f"Detected {frozen_rows} frozen header row(s) in sheet '{sheet_name}'")
                except Exception as freeze_error:
                    self.logger.debug(f"Could not check frozen panes for '{sheet_name}': {freeze_error}")
                
                # Detect if sheet has table-like structure
                try:
                    if sheet.max_row > 2 and sheet.max_column > 1:
                        metadata['has_tables'] = True
                except Exception as table_error:
                    self.logger.debug(f"Could not check table structure for '{sheet_name}': {table_error}")
            
            workbook.close()
            self.logger.debug(f"Excel analysis complete: {metadata['sheet_names']}")
            
        except Exception as e:
            self.logger.error(f"Excel analysis failed: {e}")
            import traceback
            self.logger.error(traceback.format_exc())
        
        return metadata
    
    def _is_wide_content(self, doc_obj) -> bool:
        """Determine if document has wide content requiring landscape orientation"""
        try:
            # Check if document has tables with many columns
            for item in doc_obj.iterate_items():
                if hasattr(item, 'data') and isinstance(item.data, dict):
                    if 'num_cols' in item.data and item.data['num_cols'] > 6:
                        return True
            return False
        except:
            return False
    
    def _process_docling_document(self, doc_obj, elements, styles, input_file: Path, excel_metadata: dict = None):
        """Process Docling document structure and add to PDF elements"""
        from reportlab.platypus import Table, TableStyle, Paragraph, PageBreak
        from reportlab.lib import colors
        
        if excel_metadata is None:
            excel_metadata = {}
        
        page_count = 0
        current_sheet = None
        items_processed = 0
        
        # Try different methods to extract content from Docling document
        try:
            # Method 1: Use iterate_items() if available
            if hasattr(doc_obj, 'iterate_items'):
                for item in doc_obj.iterate_items():
                    items_processed += 1
                    item_type = getattr(item, 'label', getattr(item, 'type', 'unknown'))
                    
                    # Handle different content types
                    if item_type in ('table', 'Table'):
                        page_count += 1
                        table_data = self._extract_table_data(item)
                        
                        if table_data:
                            # Add sheet header if new sheet
                            sheet_name = self._get_sheet_name(item)
                            if sheet_name and sheet_name != current_sheet:
                                current_sheet = sheet_name
                                if RAG_OPTIMIZATION_ENABLED and EXCEL_ADD_CITATION_HEADERS:
                                    elements.append(
                                        Paragraph(f"<b>Sheet: {sheet_name}</b>", styles['Heading2'])
                                    )
                            
                            # Add row/column headers if enabled (like Excel A, B, C... and 1, 2, 3...)
                            if RAG_OPTIMIZATION_ENABLED and EXCEL_PRINT_ROW_COL_HEADERS:
                                table_data = self._add_row_col_headers(table_data)
                            
                            # Get frozen panes info for this sheet
                            frozen_rows = 0
                            if sheet_name and sheet_name in excel_metadata.get('frozen_panes', {}):
                                frozen_rows = excel_metadata['frozen_panes'][sheet_name]
                            
                            # Apply table pagination if enabled
                            if RAG_OPTIMIZATION_ENABLED and EXCEL_TABLE_OPTIMIZATION and EXCEL_TABLE_MAX_ROWS_PER_PAGE > 0:
                                self._add_paginated_table(table_data, elements, input_file, frozen_rows, sheet_name)
                            else:
                                self._add_single_table(table_data, elements)
                        
                    elif item_type in ('paragraph', 'text', 'Paragraph', 'Text'):
                        text = self._extract_text(item)
                        if text and text.strip():
                            elements.append(Paragraph(text, styles['Normal']))
                    
                    elif item_type in ('heading', 'Heading', 'title', 'Title'):
                        level = getattr(item, 'level', 1)
                        style_name = f'Heading{min(level, 3)}'
                        text = self._extract_text(item)
                        if text:
                            elements.append(Paragraph(text, styles.get(style_name, styles['Heading1'])))
            
            # Method 2: Try to export as markdown and parse tables
            elif hasattr(doc_obj, 'export_to_markdown'):
                self.logger.info("Using markdown export fallback")
                md_text = doc_obj.export_to_markdown()
                if md_text:
                    items_processed += 1
                    # Simple markdown table parsing
                    elements.append(Paragraph(md_text.replace('\n', '<br/>'), styles['Normal']))
            
            # Method 3: Try to get tables directly
            elif hasattr(doc_obj, 'tables'):
                self.logger.info("Using direct table access")
                for idx, table in enumerate(doc_obj.tables):
                    items_processed += 1
                    table_data = self._extract_table_data(table)
                    if table_data:
                        if RAG_OPTIMIZATION_ENABLED and EXCEL_PRINT_ROW_COL_HEADERS:
                            table_data = self._add_row_col_headers(table_data)
                        self._add_single_table(table_data, elements)
            
            if items_processed == 0:
                self.logger.warning(f"No content extracted from Docling document. Document structure: {dir(doc_obj)[:10]}")
                # Add fallback content
                elements.append(Paragraph(f"Document: {input_file.name}", styles['Normal']))
                elements.append(Paragraph("(Content extraction failed - please check Docling version)", styles['Normal']))
                
        except Exception as e:
            self.logger.error(f"Error processing Docling document: {e}")
            elements.append(Paragraph(f"Error: {str(e)}", styles['Normal']))
    
    def _get_best_font_for_text(self, text: str) -> str:
        """Determine best font based on text content"""
        if not text:
            return self._unicode_font
            
        # Check for CJK characters
        has_japanese = any('\u3040' <= c <= '\u309f' or '\u30a0' <= c <= '\u30ff' for c in text) # Hiragana/Katakana
        has_korean = any('\uac00' <= c <= '\ud7af' for c in text) # Hangul
        has_chinese = any('\u4e00' <= c <= '\u9fff' for c in text) # Common CJK
        
        if has_japanese and self._font_map['japanese']:
            return self._font_map['japanese']
        if has_korean and self._font_map['korean']:
            return self._font_map['korean']
        if has_chinese:
            # Prefer Japanese font for Chinese if Chinese font missing (better coverage than generic)
            if self._font_map['chinese']:
                return self._font_map['chinese']
            if self._font_map['japanese']:
                return self._font_map['japanese']
                
        # For Vietnamese/Latin Extended, prefer Vietnamese-capable font
        # Check for Vietnamese specific characters (simplified check)
        has_vietnamese = any(c in 'àáạảãâầấậẩẫăằắặẳẵèéẹẻẽêềếệểễìíịỉĩòóọỏõôồốộổỗơờớợởỡùúụủũưừứựửữỳýỵỷỹđ' for c in text.lower())
        if has_vietnamese and self._font_map['vietnamese']:
            return self._font_map['vietnamese']
            
        return self._unicode_font

    def _add_excel_tables_with_openpyxl(self, input_file: Path, elements, styles, excel_metadata: dict):
        """Fallback: Extract Excel tables directly with openpyxl when Docling fails"""
        try:
            import openpyxl
            from reportlab.platypus import Paragraph, Spacer, PageBreak
            from reportlab.lib.units import inch
            
            workbook = openpyxl.load_workbook(str(input_file), data_only=True)
            sheet_names = excel_metadata.get('sheet_names', [sheet.title for sheet in workbook.worksheets])
            sheets_processed = 0
            
            self.logger.info(f"Processing {len(sheet_names)} sheets with openpyxl fallback")
            
            for idx, sheet_name in enumerate(sheet_names):
                try:
                    sheet = workbook[sheet_name]
                    
                    # Extract data from sheet
                    table_data = []
                    sheet_text_content = ""
                    
                    for row in sheet.iter_rows(values_only=True):
                        # Skip completely empty rows
                        if any(cell is not None for cell in row):
                            # Ensure proper Unicode handling
                            row_data = []
                            for cell in row:
                                if cell is None:
                                    row_data.append('')
                                else:
                                    # Convert to string and ensure Unicode
                                    cell_str = str(cell)
                                    # Handle potential encoding issues
                                    try:
                                        # Ensure it's proper Unicode
                                        cell_str.encode('utf-8').decode('utf-8')
                                        row_data.append(cell_str)
                                        sheet_text_content += cell_str
                                    except (UnicodeDecodeError, UnicodeEncodeError):
                                        # Try to fix encoding issues
                                        fixed_str = cell_str.encode('utf-8', errors='replace').decode('utf-8')
                                        row_data.append(fixed_str)
                                        sheet_text_content += fixed_str
                            table_data.append(row_data)
                    
                    # Determine best font for this sheet
                    sheet_font = self._get_best_font_for_text(sheet_text_content)
                    self.logger.debug(f"Selected font '{sheet_font}' for sheet '{sheet_name}'")
                    
                    # Add sheet header
                    if RAG_OPTIMIZATION_ENABLED and EXCEL_ADD_CITATION_HEADERS:
                        # Use sheet font for header too
                        header_style = styles['Heading2']
                        if sheet_font:
                            header_style.fontName = sheet_font
                        elements.append(Paragraph(f"<b>Sheet: {sheet_name}</b>", header_style))
                        elements.append(Spacer(1, 0.1*inch))
                    
                    if table_data:
                        # Add row/column headers if enabled
                        if RAG_OPTIMIZATION_ENABLED and EXCEL_PRINT_ROW_COL_HEADERS:
                            table_data = self._add_row_col_headers(table_data)
                        
                        # Get frozen panes info
                        frozen_rows = excel_metadata.get('frozen_panes', {}).get(sheet_name, 0)
                        
                        # Apply pagination
                        if RAG_OPTIMIZATION_ENABLED and EXCEL_TABLE_OPTIMIZATION and EXCEL_TABLE_MAX_ROWS_PER_PAGE > 0:
                            self._add_paginated_table(table_data, elements, input_file, frozen_rows, sheet_name, font_name=sheet_font)
                        else:
                            self._add_single_table(table_data, elements, font_name=sheet_font)
                        
                        sheets_processed += 1
                        self.logger.info(f"Processed sheet {idx+1}/{len(sheet_names)}: {sheet_name} ({len(table_data)} rows)")
                    else:
                        self.logger.warning(f"Sheet '{sheet_name}' has no data, skipping")
                    
                    # Add page break between sheets (except for last sheet)
                    if idx < len(sheet_names) - 1:
                        elements.append(PageBreak())
                    else:
                        elements.append(Spacer(1, 0.2*inch))
                        
                except Exception as sheet_error:
                    self.logger.error(f"Failed to process sheet '{sheet_name}': {sheet_error}")
            
            workbook.close()
            self.logger.info(f"Successfully extracted {sheets_processed} sheets using openpyxl fallback")
            
        except Exception as e:
            self.logger.error(f"openpyxl fallback failed: {e}")
            import traceback
            self.logger.debug(traceback.format_exc())
    
    def _get_sheet_name(self, item) -> str:
        """Extract sheet name from Docling item"""
        try:
            if hasattr(item, 'prov') and item.prov:
                return getattr(item.prov, 'sheet_name', None) or getattr(item.prov, 'page_no', None)
            if hasattr(item, 'metadata') and isinstance(item.metadata, dict):
                return item.metadata.get('sheet_name') or item.metadata.get('source')
        except:
            pass
        return None
    
    def _extract_text(self, item) -> str:
        """Extract text from Docling item"""
        try:
            if hasattr(item, 'text'):
                return str(item.text)
            elif hasattr(item, 'content'):
                return str(item.content)
            elif hasattr(item, '__str__'):
                return str(item)
        except:
            pass
        return ""
    
    def _extract_table_data(self, table_item):
        """Extract table data from Docling table item"""
        try:
            # Method 1: Direct data attribute
            if hasattr(table_item, 'data'):
                if isinstance(table_item.data, list) and table_item.data:
                    return table_item.data
                elif isinstance(table_item.data, dict):
                    # Try to extract from dict structure
                    if 'grid' in table_item.data:
                        return table_item.data['grid']
                    elif 'rows' in table_item.data:
                        return table_item.data['rows']
            
            # Method 2: Export to DataFrame
            if hasattr(table_item, 'export_to_dataframe'):
                df = table_item.export_to_dataframe()
                if df is not None and not df.empty:
                    # Include column headers as first row
                    return [df.columns.tolist()] + df.values.tolist()
            
            # Method 3: Cells grid
            if hasattr(table_item, 'cells') and table_item.cells:
                return self._cells_to_grid(table_item.cells)
            
            # Method 4: Try export_to_html and parse
            if hasattr(table_item, 'export_to_html'):
                html = table_item.export_to_html()
                if html:
                    self.logger.debug(f"Table extracted as HTML (length: {len(html)})")
                    # For now, return indication that we have table content
                    return [["Table content available in HTML format"]]
            
            # Method 5: Direct text extraction
            if hasattr(table_item, 'text'):
                text = str(table_item.text)
                if text.strip():
                    # Try to parse as simple rows
                    lines = text.strip().split('\n')
                    if lines:
                        return [line.split('\t') for line in lines]
            
            self.logger.debug(f"No table data extracted. Available attributes: {dir(table_item)[:20]}")
            return None
            
        except Exception as e:
            self.logger.debug(f"Failed to extract table data: {e}")
            import traceback
            self.logger.debug(traceback.format_exc())
            return None
    
    def _cells_to_grid(self, cells) -> list:
        """Convert cell structure to 2D grid"""
        # This is a simplified implementation
        # Real implementation would need to handle merged cells
        max_row = max(cell.row for cell in cells) + 1
        max_col = max(cell.col for cell in cells) + 1
        
        grid = [['' for _ in range(max_col)] for _ in range(max_row)]
        
        for cell in cells:
            grid[cell.row][cell.col] = str(cell.text)
        
        return grid
    
    def _add_row_col_headers(self, table_data: list) -> list:
        """Add Excel-style row/column headers (A, B, C... and 1, 2, 3...)"""
        if not table_data:
            return table_data
        
        # Generate column headers (A, B, C, ...)
        def col_num_to_letter(n):
            """Convert column number to Excel-style letter (1=A, 27=AA)"""
            result = ""
            while n > 0:
                n -= 1
                result = chr(65 + (n % 26)) + result
                n //= 26
            return result
        
        num_cols = len(table_data[0]) if table_data else 0
        col_headers = [''] + [col_num_to_letter(i+1) for i in range(num_cols)]
        
        # Add column headers as first row
        new_data = [col_headers]
        
        # Add row numbers and data
        for i, row in enumerate(table_data, start=1):
            new_data.append([str(i)] + list(row))
        
        return new_data
    
    def _add_paginated_table(self, table_data, elements, input_file: Path, frozen_rows: int = 0, sheet_name: str = None, font_name: str = None):
        """Add table with pagination based on EXCEL_TABLE_MAX_ROWS_PER_PAGE"""
        from reportlab.platypus import Table, TableStyle, PageBreak, Paragraph
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet
        
        if not table_data or len(table_data) < 2:
            self._add_single_table(table_data, elements, font_name)
            return
        
        # Determine header rows (frozen panes take precedence)
        header_row_count = max(frozen_rows, 1)
        if header_row_count >= len(table_data):
            header_row_count = 1
        
        headers = table_data[:header_row_count]
        data_rows = table_data[header_row_count:]
        max_rows = EXCEL_TABLE_MAX_ROWS_PER_PAGE
        
        if frozen_rows > 0:
            self.logger.info(f"Repeating {frozen_rows} frozen header row(s) on each page")
        
        # Split into chunks
        total_pages = (len(data_rows) + max_rows - 1) // max_rows
        styles = getSampleStyleSheet()
        
        for page_num, i in enumerate(range(0, len(data_rows), max_rows), start=1):
            chunk = data_rows[i:i + max_rows]
            chunk_data = list(headers) + chunk  # Repeat headers on each page
            
            # Add page citation for multi-page tables
            if RAG_OPTIMIZATION_ENABLED and EXCEL_ADD_CITATION_HEADERS and total_pages > 1:
                citation = f"Table page {page_num} of {total_pages}"
                if sheet_name:
                    citation += f" (Sheet: {sheet_name})"
                elements.append(Paragraph(f"<i>{citation}</i>", styles['Normal']))
            
            table = Table(chunk_data)
            self._apply_table_style(table, header_row_count, font_name)
            elements.append(table)
            
            if i + max_rows < len(data_rows):
                elements.append(PageBreak())
    
    def _add_single_table(self, table_data, elements, font_name: str = None):
        """Add table without pagination"""
        from reportlab.platypus import Table
        
        if not table_data:
            return
        
        table = Table(table_data)
        self._apply_table_style(table, font_name=font_name)
        elements.append(table)
    
    def _apply_table_style(self, table, header_rows: int = 1, font_name: str = None):
        """Apply RAG-optimized table styling"""
        from reportlab.platypus import TableStyle
        from reportlab.lib import colors
        
        # Style header rows
        # Use Unicode fonts if available for CJK support
        # If specific font provided (per sheet), use it. Otherwise use global default.
        header_font = font_name if font_name else (self._unicode_font_bold if self._unicode_font_bold else 'Helvetica-Bold')
        body_font = font_name if font_name else (self._unicode_font if self._unicode_font else 'Helvetica')
        
        style_commands = [
            ('BACKGROUND', (0, 0), (-1, header_rows - 1), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, header_rows - 1), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, header_rows - 1), header_font),
            ('FONTSIZE', (0, 0), (-1, header_rows - 1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, header_rows - 1), 12),
            ('BACKGROUND', (0, header_rows), (-1, -1), colors.beige),
            ('FONTNAME', (0, header_rows), (-1, -1), body_font),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
        ]
        
        # Add gridlines if enabled (RAG optimization for better structure recognition)
        if RAG_OPTIMIZATION_ENABLED and EXCEL_PRINT_GRIDLINES:
            style_commands.append(('GRID', (0, 0), (-1, -1), 0.5, colors.black))
        else:
            # At least add lines around headers
            style_commands.append(('LINEBELOW', (0, header_rows - 1), (-1, header_rows - 1), 1, colors.black))
        
        table.setStyle(TableStyle(style_commands))
    
    def _add_page_header_footer(self, canvas, doc):
        """Add citation headers/footers to each page (RAG optimization)"""
        if not RAG_OPTIMIZATION_ENABLED or not EXCEL_ADD_CITATION_HEADERS:
            return
        
        from reportlab.lib.units import inch
        
        # Save state
        canvas.saveState()
        
        # Add page number footer
        page_num = canvas.getPageNumber()
        footer_text = f"Page {page_num}"
        footer_font = self._unicode_font if self._unicode_font else 'Helvetica'
        canvas.setFont(footer_font, 9)
        canvas.drawCentredString(doc.pagesize[0] / 2, 0.5 * inch, footer_text)
        
        # Restore state
        canvas.restoreState()
    
    def _apply_pdf_metadata(self, pdf_path: Path, input_file: Path, doc_obj, excel_metadata: dict = None):
        """Apply PDF metadata for RAG optimization"""
        if not RAG_OPTIMIZATION_ENABLED:
            return
        
        if excel_metadata is None:
            excel_metadata = {}
        
        try:
            from pypdf import PdfReader, PdfWriter
        except ImportError:
            self.logger.debug("pypdf not installed; skipping metadata")
            return
        
        try:
            reader = PdfReader(str(pdf_path))
            writer = PdfWriter()
            for page in reader.pages:
                writer.add_page(page)
            
            page_count = len(reader.pages)
            metadata_parts = []
            keywords = [input_file.stem, 'Docling', 'RAG']
            
            # Add Excel-specific metadata
            sheet_count = excel_metadata.get('sheet_count', 0)
            sheet_names = excel_metadata.get('sheet_names', [])
            
            if CITATION_INCLUDE_FILENAME:
                metadata_parts.append(f"Source: {input_file.name}")
                keywords.append(input_file.name)
            if sheet_count > 0:
                metadata_parts.append(f"Sheets: {sheet_count}")
                keywords.append(f"sheets:{sheet_count}")
                if sheet_names:
                    keywords.extend(sheet_names[:3])  # Add first 3 sheet names
            if CITATION_INCLUDE_PAGE and page_count:
                metadata_parts.append(f"Pages: {page_count}")
                keywords.append(f"pages:{page_count}")
            if CITATION_INCLUDE_DATE:
                metadata_parts.append(
                    f"Converted: {datetime.now().strftime(CITATION_DATE_FORMAT)}"
                )
            
            metadata = {
                '/Title': input_file.stem,
                '/Author': 'AI4Team Docling Converter',
                '/Subject': 'RAG Export - Enhanced Layout',
                '/Creator': 'Docling',
                '/Producer': 'ReportLab + Docling',
                '/Keywords': ', '.join(dict.fromkeys(keywords)),
            }
            if metadata_parts:
                metadata['/Comments'] = ' | '.join(metadata_parts)
            
            existing_metadata = dict(reader.metadata or {})
            existing_metadata.update(metadata)
            writer.add_metadata(existing_metadata)
            
            temp_path = pdf_path.parent / (pdf_path.name + '.tmp')
            with open(temp_path, 'wb') as f:
                writer.write(f)
            temp_path.replace(pdf_path)
            
        except Exception as e:
            self.logger.debug(f"Failed to apply metadata: {e}")
