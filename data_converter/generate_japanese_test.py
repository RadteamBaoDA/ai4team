"""
Generate test Excel file with Japanese and other CJK characters
to verify Unicode support in Docling converter
"""

import openpyxl
from pathlib import Path

def create_japanese_test_excel():
    """Create Excel with Japanese content"""
    wb = openpyxl.Workbook()
    
    # Sheet 1: Japanese
    ws1 = wb.active
    ws1.title = "日本語テスト"
    
    ws1['A1'] = '名前'
    ws1['B1'] = '年齢'
    ws1['C1'] = '都市'
    ws1['D1'] = '職業'
    
    ws1['A2'] = '田中太郎'
    ws1['B2'] = 25
    ws1['C2'] = '東京'
    ws1['D2'] = 'エンジニア'
    
    ws1['A3'] = '鈴木花子'
    ws1['B3'] = 30
    ws1['C3'] = '大阪'
    ws1['D3'] = 'デザイナー'
    
    ws1['A4'] = '佐藤次郎'
    ws1['B4'] = 28
    ws1['C4'] = '京都'
    ws1['D4'] = 'プログラマー'
    
    # Add some summary text
    ws1['A6'] = '概要:'
    ws1['A7'] = 'これは日本語のテストファイルです。'
    ws1['A8'] = 'PDFに正しく変換されることを確認します。'
    
    # Sheet 2: Mixed languages
    ws2 = wb.create_sheet("多言語")
    
    ws2['A1'] = '言語'
    ws2['B1'] = 'テキスト'
    ws2['C1'] = '説明'
    
    ws2['A2'] = '日本語'
    ws2['B2'] = 'こんにちは世界'
    ws2['C2'] = 'ひらがな、カタカナ、漢字'
    
    ws2['A3'] = '中国語'
    ws2['B3'] = '你好世界'
    ws2['C3'] = '简体中文'
    
    ws2['A4'] = '韓国語'
    ws2['B4'] = '안녕하세요'
    ws2['C4'] = '한글'
    
    ws2['A5'] = 'English'
    ws2['B5'] = 'Hello World'
    ws2['C5'] = 'ASCII characters'
    
    ws2['A6'] = '特殊文字'
    ws2['B6'] = '①②③④⑤ ★☆♪♫'
    ws2['C6'] = '記号とマーク'
    
    # Sheet 3: Business document in Japanese
    ws3 = wb.create_sheet("売上報告")
    
    ws3['A1'] = '月'
    ws3['B1'] = '売上（円）'
    ws3['C1'] = '利益（円）'
    ws3['D1'] = '成長率'
    
    ws3['A2'] = '1月'
    ws3['B2'] = 1000000
    ws3['C2'] = 200000
    ws3['D2'] = '5%'
    
    ws3['A3'] = '2月'
    ws3['B3'] = 1200000
    ws3['C3'] = 250000
    ws3['D3'] = '20%'
    
    ws3['A4'] = '3月'
    ws3['B4'] = 1500000
    ws3['C4'] = 350000
    ws3['D4'] = '25%'
    
    ws3['A6'] = '備考:'
    ws3['A7'] = '第一四半期の売上は順調に増加しています。'
    ws3['A8'] = '来月も引き続き成長を目指します。'
    
    # Save
    output_dir = Path(__file__).parent / 'input'
    output_dir.mkdir(exist_ok=True)
    
    output_path = output_dir / 'japanese_test.xlsx'
    wb.save(output_path)
    print(f"Created Japanese test Excel: {output_path}")
    print(f"Sheets: {[ws.title for ws in wb.worksheets]}")
    
    return output_path

if __name__ == '__main__':
    create_japanese_test_excel()
