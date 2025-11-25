#!/usr/bin/env python3
"""
Generate sample Excel files for testing Docling converter.

This script creates complex Excel files with various table structures
to test the enhanced layout recognition capabilities of Docling.

Usage:
    python generate_test_excel.py
"""

import os
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("❌ openpyxl not installed")
    print("\nInstall with:")
    print("  pip install openpyxl")
    sys.exit(1)


def create_simple_table(output_dir: Path):
    """Create a simple Excel table."""
    print("\n📄 Creating simple_table.xlsx...")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Data"
    
    # Headers
    headers = ["Product", "Q1", "Q2", "Q3", "Q4", "Total"]
    ws.append(headers)
    
    # Style headers
    for col in range(1, len(headers) + 1):
        cell = ws.cell(1, col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Data
    data = [
        ["Laptop", 1200, 1350, 1500, 1600, "=SUM(B2:E2)"],
        ["Mouse", 450, 500, 520, 480, "=SUM(B3:E3)"],
        ["Keyboard", 380, 420, 400, 450, "=SUM(B4:E4)"],
        ["Monitor", 890, 950, 1000, 1100, "=SUM(B5:E5)"],
    ]
    
    for row in data:
        ws.append(row)
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    # Auto-width columns
    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2
    
    output_path = output_dir / "simple_table.xlsx"
    wb.save(output_path)
    print(f"✅ Created: {output_path}")


def create_complex_merged_cells(output_dir: Path):
    """Create Excel with complex merged cells."""
    print("\n📄 Creating complex_merged.xlsx...")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Quarterly Report"
    
    # Title (merged)
    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value = "2024 Quarterly Sales Report"
    title_cell.font = Font(size=16, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30
    
    # Headers with merged cells
    ws.merge_cells("A2:A3")
    ws["A2"].value = "Region"
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    
    ws.merge_cells("B2:C2")
    ws["B2"].value = "Q1"
    ws["B2"].alignment = Alignment(horizontal="center")
    ws["B3"].value = "Sales"
    ws["C3"].value = "Growth %"
    
    ws.merge_cells("D2:E2")
    ws["D2"].value = "Q2"
    ws["D2"].alignment = Alignment(horizontal="center")
    ws["D3"].value = "Sales"
    ws["E3"].value = "Growth %"
    
    ws.merge_cells("F2:F3")
    ws["F2"].value = "Total"
    ws["F2"].alignment = Alignment(horizontal="center", vertical="center")
    
    # Style headers
    header_fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
    for row in [2, 3]:
        for col in range(1, 7):
            cell = ws.cell(row, col)
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Data
    data = [
        ["North", 45000, 12, 52000, 15, "=B4+D4"],
        ["South", 38000, 8, 41000, 7, "=B5+D5"],
        ["East", 51000, 18, 60000, 17, "=B6+D6"],
        ["West", 42000, 10, 46000, 9, "=B7+D7"],
    ]
    
    for row in data:
        ws.append(row)
    
    # Borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in ws.iter_rows(min_row=2, max_row=7, min_col=1, max_col=6):
        for cell in row:
            cell.border = thin_border
    
    # Auto-width
    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    output_path = output_dir / "complex_merged.xlsx"
    wb.save(output_path)
    print(f"✅ Created: {output_path}")


def create_multi_sheet_workbook(output_dir: Path):
    """Create workbook with multiple sheets."""
    print("\n📄 Creating multi_sheet.xlsx...")
    
    wb = Workbook()
    
    # Sheet 1: Summary
    ws1 = wb.active
    ws1.title = "Summary"
    ws1["A1"] = "Overall Summary"
    ws1["A1"].font = Font(size=14, bold=True)
    ws1["A3"] = "Total Revenue"
    ws1["B3"] = 500000
    ws1["A4"] = "Total Expenses"
    ws1["B4"] = 350000
    ws1["A5"] = "Net Profit"
    ws1["B5"] = "=B3-B4"
    
    # Sheet 2: Monthly Data
    ws2 = wb.create_sheet("Monthly Data")
    headers = ["Month", "Revenue", "Expenses", "Profit"]
    ws2.append(headers)
    
    months_data = [
        ["January", 45000, 32000, "=B2-C2"],
        ["February", 42000, 30000, "=B3-C3"],
        ["March", 48000, 33000, "=B4-C4"],
        ["April", 50000, 34000, "=B5-C5"],
    ]
    
    for row in months_data:
        ws2.append(row)
    
    # Style headers
    for col in range(1, 5):
        cell = ws2.cell(1, col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    
    # Sheet 3: Products
    ws3 = wb.create_sheet("Products")
    ws3["A1"] = "Product Inventory"
    ws3["A1"].font = Font(size=12, bold=True)
    
    product_headers = ["SKU", "Name", "Stock", "Price", "Value"]
    ws3.append(product_headers)
    
    products = [
        ["LAP001", "Business Laptop", 45, 1200, "=C3*D3"],
        ["MOU001", "Wireless Mouse", 230, 25, "=C4*D4"],
        ["KEY001", "Mechanical Keyboard", 120, 80, "=C5*D5"],
    ]
    
    for row in products:
        ws3.append(row)
    
    output_path = output_dir / "multi_sheet.xlsx"
    wb.save(output_path)
    print(f"✅ Created: {output_path}")


def create_large_table_pagination(output_dir: Path):
    """Create Excel with large table to test pagination."""
    print("\n📄 Creating large_table.xlsx...")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Large Dataset"
    
    # Headers
    headers = ["ID", "Name", "Department", "Salary", "Start Date", "Status"]
    ws.append(headers)
    
    # Style headers
    for col in range(1, len(headers) + 1):
        cell = ws.cell(1, col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    
    # Generate 50 rows (to test pagination with default 15 rows/page)
    departments = ["Sales", "Engineering", "Marketing", "HR", "Finance"]
    statuses = ["Active", "On Leave", "Remote"]
    
    for i in range(1, 51):
        row = [
            f"EMP{i:04d}",
            f"Employee {i}",
            departments[i % len(departments)],
            50000 + (i * 1000),
            f"2024-{(i % 12) + 1:02d}-01",
            statuses[i % len(statuses)]
        ]
        ws.append(row)
    
    # Freeze header
    ws.freeze_panes = "A2"
    
    # Auto-width
    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_length + 2, 20)
    
    output_path = output_dir / "large_table.xlsx"
    wb.save(output_path)
    print(f"✅ Created: {output_path}")


def create_borderless_table(output_dir: Path):
    """Create Excel with borderless table (tests layout detection)."""
    print("\n📄 Creating borderless_table.xlsx...")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Borderless"
    
    # Data without explicit borders (tests Docling's table detection)
    data = [
        ["Category", "Item", "Quantity", "Unit Price", "Total"],
        ["Office", "Pens", 100, 0.5, "=C2*D2"],
        ["Office", "Notebooks", 50, 2.0, "=C3*D3"],
        ["IT", "USB Drives", 25, 15, "=C4*D4"],
        ["IT", "Monitors", 10, 250, "=C5*D5"],
    ]
    
    for row in data:
        ws.append(row)
    
    # Only style headers (no borders)
    for col in range(1, 6):
        cell = ws.cell(1, col)
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal="left")
    
    # Add some spacing
    for col in ws.columns:
        ws.column_dimensions[get_column_letter(col[0].column)].width = 15
    
    output_path = output_dir / "borderless_table.xlsx"
    wb.save(output_path)
    print(f"✅ Created: {output_path}")


def main():
    """Generate all test Excel files."""
    print("=" * 60)
    print("  Generating Test Excel Files for Docling")
    print("=" * 60)
    
    # Create test directory
    test_dir = Path(__file__).parent / "input" / "docling_test"
    test_dir.mkdir(parents=True, exist_ok=True)
    
    print(f"\nOutput directory: {test_dir}")
    
    # Generate all test files
    create_simple_table(test_dir)
    create_complex_merged_cells(test_dir)
    create_multi_sheet_workbook(test_dir)
    create_large_table_pagination(test_dir)
    create_borderless_table(test_dir)
    
    print("\n" + "=" * 60)
    print("✅ All test files generated!")
    print("=" * 60)
    
    print("\nTest files created:")
    print(f"  📁 {test_dir}")
    print("  - simple_table.xlsx       (basic table with formulas)")
    print("  - complex_merged.xlsx     (merged cells, complex headers)")
    print("  - multi_sheet.xlsx        (multiple worksheets)")
    print("  - large_table.xlsx        (50 rows, tests pagination)")
    print("  - borderless_table.xlsx   (no borders, tests detection)")
    
    print("\nNext steps:")
    print("  1. Install Docling: pip install docling")
    print("  2. Enable Docling: export USE_DOCLING_CONVERTER=true")
    print(f"  3. Convert: python main.py {test_dir} ./output/docling_test")
    print("  4. Compare with standard converters")


if __name__ == "__main__":
    main()
